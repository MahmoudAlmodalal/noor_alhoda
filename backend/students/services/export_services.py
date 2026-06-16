"""
Excel export for the students list.

Builds an xlsx file with the common student fields plus a computed
memorization-parts count. Called from ``StudentExportApi``; the selector
layer (``student_list``) handles row-level RBAC so the view passes a
pre-filtered QuerySet through.
"""
from io import BytesIO

from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADERS = [
    "الاسم الرباعي",
    "رقم الهوية",
    "تاريخ الميلاد",
    "الصف الدراسي",
    "رقم الجوال",
    "العنوان",
    "اسم ولي الأمر",
    "رقم جوال ولي الأمر",
    "المحفظ",
    "اسم الحلقة",
    "الحالة الصحية",
    "المهارات",
    "عدد الأجزاء المحفوظة",
    "تاريخ التسجيل",
]


_HEALTH_LABELS = {
    "normal": "طبيعي",
    "martyr_son": "ابن شهيد",
    "injured": "جريح",
    "sick": "مريض",
    "other": "أخرى",
}


_SKILL_LABELS = {
    "quran": "قراءات قرآن",
    "nasheed": "إنشاد",
    "poetry": "شعر",
    "other": "أخرى",
}


def _skills_to_ar(skills) -> str:
    if not isinstance(skills, dict) or not skills:
        return "—"
    picked = [_SKILL_LABELS[k] for k, v in skills.items() if k in _SKILL_LABELS and v is True]
    return "، ".join(picked) if picked else "—"


def _memorized_ajza_for(student) -> int:
    """Rough heuristic matching ``student_stats``: 604 pages across 30 juz."""
    from records.models import WeeklyPlan

    total = (
        WeeklyPlan.objects.filter(student=student).aggregate(
            total=Sum("total_achieved")
        )["total"]
        or 0
    )
    pages_per_juz = 604 // 30
    return total // pages_per_juz if total > 0 else 0


def generate_students_xlsx(students) -> bytes:
    """Return the rendered xlsx file as bytes for the given student iterable."""

    wb = Workbook()
    ws = wb.active
    ws.title = "الطلاب"
    ws.sheet_view.rightToLeft = True

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="0B5394")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, title in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, student in enumerate(students, start=2):
        teacher = getattr(student, "teacher", None)
        national_id = getattr(getattr(student, "user", None), "national_id", "") or ""
        values = [
            student.full_name,
            national_id,
            student.birthdate.isoformat() if student.birthdate else "",
            student.grade,
            student.mobile or "",
            student.address or "",
            student.guardian_name or "",
            student.guardian_mobile or "",
            teacher.full_name if teacher else "غير مسنَد",
            (teacher.ring_name if teacher else "") or "",
            _HEALTH_LABELS.get(student.health_status, student.health_status or "—"),
            _skills_to_ar(student.skills),
            _memorized_ajza_for(student),
            student.enrollment_date.isoformat() if student.enrollment_date else "",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col, value=value)

    ws.row_dimensions[1].height = 28
    for col in range(1, len(_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
