"""
Excel export for the teachers list, plus an empty template generator.

Mirrors ``students/services/export_services.py``: openpyxl with RTL
layout, blue/white header styling, column widths matching the student
sheet. The column order comes from ``excel_field_map.HEADERS`` so that
the export, the template, and the import all agree.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from teacher.services.excel_field_map import (
    HEADERS,
    affiliation_to_arabic,
    job_title_to_arabic,
    marital_to_arabic,
    session_days_to_arabic,
)


_HEADER_FILL_HEX = "0B5394"


def _style_header_row(ws) -> None:
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL_HEX)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    ws.row_dimensions[1].height = 28
    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22


def _row_for_teacher(teacher) -> list:
    user = getattr(teacher, "user", None)
    national_id = getattr(user, "national_id", "") or ""
    phone = getattr(user, "phone_number", "") or ""

    courses_text = "، ".join(c.name for c in teacher.courses.all())

    return [
        teacher.full_name or "",
        national_id,
        phone,
        teacher.birthdate.isoformat() if teacher.birthdate else "",
        marital_to_arabic(teacher.marital_status),
        teacher.education_qualification or "",
        teacher.specialization or "",
        teacher.last_tajweed_course or "",
        job_title_to_arabic(teacher.job_title),
        affiliation_to_arabic(teacher.affiliation),
        teacher.ring_name or "",
        session_days_to_arabic(teacher.session_days or []),
        teacher.max_students,
        courses_text,
        teacher.family_members_count if teacher.family_members_count is not None else "",
        teacher.wallet_name or "",
        teacher.wallet_number or "",
        teacher.created_at.isoformat() if teacher.created_at else "",
        teacher.updated_at.isoformat() if teacher.updated_at else "",
    ]


def generate_teachers_xlsx(teachers) -> bytes:
    """Return the rendered xlsx file as bytes for the given teacher iterable."""
    wb = Workbook()
    ws = wb.active
    ws.title = "المحفظون"
    ws.sheet_view.rightToLeft = True

    _style_header_row(ws)

    for row_idx, teacher in enumerate(teachers, start=2):
        for col, value in enumerate(_row_for_teacher(teacher), start=1):
            ws.cell(row=row_idx, column=col, value=value)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_teachers_template_xlsx() -> bytes:
    """Return an empty template — headers + one example row.

    The example row uses Arabic enum labels so an admin can copy the
    formatting and replace values without guessing.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "المحفظون"
    ws.sheet_view.rightToLeft = True

    _style_header_row(ws)

    example_row = [
        "أحمد محمد عبدالله الخطيب",       # full_name
        "401234567",                         # national_id
        "0599123456",                        # phone_number
        "1990-05-12",                        # birthdate
        "متزوج",                             # marital_status
        "بكالوريوس شريعة",                   # education_qualification
        "إجازة بالقراءات السبع",             # specialization
        "دورة تجويد متقدمة 2024",           # last_tajweed_course
        "محفظ",                              # job_title
        "دار القرآن",                        # affiliation
        "حلقة الفجر",                        # ring_name
        "السبت، الأحد، الاثنين",             # session_days
        25,                                   # max_students
        "حفظ القرآن، تجويد",                # courses (names only)
        4,                                    # family_members_count
        "بنك فلسطين",                        # wallet_name
        "9701234567",                        # wallet_number
        # created_at / updated_at are export-only — leave blank in the
        # template so admins know they don't fill them.
        "",
        "",
    ]
    for col, value in enumerate(example_row, start=1):
        ws.cell(row=2, column=col, value=value)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
