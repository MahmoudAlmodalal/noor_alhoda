"""
xlsx export, template download, and file-upload import for teachers.

Mirrors ``students/views/student_export_views.py`` for the export, and
parses the upload here (in the view) before handing pre-shaped row dicts
to ``teacher_excel_bulk_import``. Header matching is tolerant of NBSP /
RLM / LRM marks because Arabic xlsx files round-trip through them.
"""
from datetime import date

from django.http import HttpResponse
from openpyxl import load_workbook
from rest_framework import serializers, status
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from core.permissions import IsAdmin
from students.services.excel_import.normalization import _normalize_arabic_key
from teacher.selectors.teacher_selectors import teacher_list
from teacher.services.excel_field_map import NORMALIZED_HEADER_TO_FIELD
from teacher.services.excel_import.teacher_orchestrator import (
    teacher_excel_bulk_import,
)
from teacher.services.export_services import (
    generate_teachers_template_xlsx,
    generate_teachers_xlsx,
)


_XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def _xlsx_response(content: bytes, filename: str) -> HttpResponse:
    response = HttpResponse(content, content_type=_XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------
class TeacherExportApi(APIView):
    """GET /api/users/teachers/export/ — تصدير قائمة المحفظين إلى Excel."""

    permission_classes = [IsAdmin]

    def get(self, request):
        teachers = teacher_list(filters={})
        content = generate_teachers_xlsx(teachers)
        filename = f"teachers-{date.today():%Y-%m-%d}.xlsx"
        return _xlsx_response(content, filename)


# ---------------------------------------------------------------------------
# Template
# ---------------------------------------------------------------------------
class TeacherTemplateApi(APIView):
    """GET /api/users/teachers/template/ — تحميل قالب فارغ للاستيراد."""

    permission_classes = [IsAdmin]

    def get(self, request):
        content = generate_teachers_template_xlsx()
        return _xlsx_response(content, "teachers-template.xlsx")


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------
class _ImportSerializer(serializers.Serializer):
    file = serializers.FileField()

    def validate_file(self, value):
        name = (getattr(value, "name", "") or "").lower()
        if not name.endswith(".xlsx"):
            raise serializers.ValidationError(
                "يجب أن يكون الملف بصيغة .xlsx."
            )
        return value


def _row_is_empty(row_values) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in row_values)


def _parse_xlsx(file_obj) -> list[dict]:
    """Read the active sheet and return a list of dicts keyed by internal field.

    Header matching uses ``_normalize_arabic_key`` so NBSP, RLM/LRM, and
    trailing spaces don't silently drop columns.
    """
    wb = load_workbook(file_obj, data_only=True, read_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []

    column_to_field: dict[int, str] = {}
    for idx, header in enumerate(header_row):
        if header is None:
            continue
        key = _normalize_arabic_key(str(header))
        field = NORMALIZED_HEADER_TO_FIELD.get(key)
        if field is not None:
            column_to_field[idx] = field

    if not column_to_field:
        return []

    parsed: list[dict] = []
    for raw_row in rows_iter:
        if raw_row is None or _row_is_empty(raw_row):
            continue
        row_dict: dict = {}
        for col_idx, field in column_to_field.items():
            if col_idx < len(raw_row):
                row_dict[field] = raw_row[col_idx]
        if row_dict:
            parsed.append(row_dict)
    return parsed


class TeacherImportXlsxApi(APIView):
    """POST /api/users/teachers/import-xlsx/ — استيراد المحفظين من Excel."""

    permission_classes = [IsAdmin]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        serializer = _ImportSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(
                {"success": False, "errors": serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )

        upload = serializer.validated_data["file"]

        try:
            rows = _parse_xlsx(upload)
        except Exception:
            return Response(
                {
                    "success": False,
                    "errors": {
                        "file": "تعذّر قراءة الملف. تأكد من أنه ملف Excel صالح."
                    },
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not rows:
            return Response(
                {
                    "success": False,
                    "errors": {
                        "file": "لم يتم العثور على صفوف صالحة. تحقق من رؤوس الأعمدة."
                    },
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        result = teacher_excel_bulk_import(creator=request.user, rows=rows)
        return Response(
            {"success": True, "data": result},
            status=status.HTTP_200_OK,
        )
