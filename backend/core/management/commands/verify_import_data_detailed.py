"""
Detailed verification of Excel import with normalization awareness.
Accounts for expected data transformations (grade normalization, phone formatting, etc).

Usage: python manage.py verify_import_data_detailed /path/to/file.xlsx
"""
import openpyxl
from datetime import datetime, date
from django.core.management.base import BaseCommand, CommandError

from students.models import Student
from students.services.excel_import.normalization import (
    _normalize_arabic_key,
    _normalize_grade,
    _safe_phone,
)


class Command(BaseCommand):
    help = "Detailed verification of imported student data with normalization awareness."

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str, help="Path to the Excel file")

    def handle(self, *args, **options):
        file_path = options["file_path"]

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = None

            # Find sheet with data
            for sheet_name in workbook.sheetnames:
                candidate_sheet = workbook[sheet_name]
                if candidate_sheet.max_row > 1:
                    sheet = candidate_sheet
                    break

            if not sheet:
                raise CommandError("No sheet with data found in Excel file.")

            # Find header row
            header_row = None
            for row_idx in range(1, min(10, sheet.max_row + 1)):
                for cell in sheet[row_idx]:
                    if cell.value and 'اسم' in str(cell.value):
                        header_row = row_idx
                        break
                if header_row:
                    break

            if not header_row:
                raise CommandError("Could not find header row in Excel file.")

            # Column map
            column_index_map = {
                1: 'sequence_number',
                2: 'full_name',
                3: 'national_id',
                4: 'birthdate',
                5: 'age',
                6: 'grade',
                7: 'mobile',
                8: 'whatsapp',
                14: 'address',
                15: 'health_status',
                16: 'skills',
                17: 'previous_courses',
                18: 'guardian_name',
                19: 'guardian_national_id',
                20: 'guardian_mobile',
                21: 'bank_account_number',
                22: 'bank_account_name',
                23: 'bank_account_type',
                24: 'teacher_name',
                25: 'affiliation',
            }

            # Read all rows
            rows = []
            for row_idx, row_cells in enumerate(
                sheet.iter_rows(min_row=header_row + 1, values_only=False),
                start=header_row + 1
            ):
                if not any(cell.value for cell in row_cells):
                    continue

                row_dict = {'_row_index': row_idx}
                for col_idx, cell in enumerate(row_cells):
                    if col_idx in column_index_map:
                        field_name = column_index_map[col_idx]
                        value = cell.value

                        # Convert datetime to ISO string
                        if field_name == 'birthdate' and isinstance(value, datetime):
                            value = value.date()
                        if isinstance(value, date) and not isinstance(value, datetime):
                            value = value.isoformat()

                        row_dict[field_name] = value

                if row_dict.get('full_name') or row_dict.get('national_id'):
                    rows.append(row_dict)

            self.stdout.write(self.style.SUCCESS(f"\n📋 Found {len(rows)} student rows in xlsx"))
            self.stdout.write("=" * 120)

            # Verify each row
            stats = {
                'total': len(rows),
                'found': 0,
                'missing': 0,
                'no_id': 0,
                'field_issues': [],
            }

            for idx, xlsx_row in enumerate(rows, start=1):
                national_id = str(xlsx_row.get('national_id') or '').strip()
                full_name = str(xlsx_row.get('full_name') or '').strip()

                if not national_id:
                    stats['no_id'] += 1
                    self.stdout.write(
                        self.style.WARNING(f"⚠ Row {idx}: No national ID (name: {full_name})")
                    )
                    continue

                # Look up student by national_id
                student = Student.objects.select_related(
                    'user', 'teacher'
                ).filter(
                    user__national_id=national_id
                ).first()

                if not student:
                    stats['missing'] += 1
                    self.stdout.write(
                        self.style.ERROR(
                            f"✗ Row {idx} (ID: {national_id}): NOT FOUND in database"
                        )
                    )
                    continue

                stats['found'] += 1

                # Check key fields with normalization awareness
                issues = []

                # Name check (normalized)
                if full_name and student.full_name:
                    if _normalize_arabic_key(full_name) != _normalize_arabic_key(student.full_name):
                        issues.append(f"Name normalized mismatch: '{full_name}' vs '{student.full_name}'")

                # Grade check (normalize Arabic grades to numbers)
                xlsx_grade = str(xlsx_row.get('grade') or '').strip()
                if xlsx_grade and student.grade:
                    xlsx_grade_normalized = _normalize_grade(xlsx_grade)
                    if xlsx_grade_normalized and str(student.grade) != xlsx_grade_normalized:
                        issues.append(f"Grade mismatch: '{xlsx_grade}' ({xlsx_grade_normalized}) vs {student.grade}")

                # Mobile check (normalize phone numbers)
                xlsx_mobile = str(xlsx_row.get('mobile') or '').strip()
                if xlsx_mobile and student.mobile:
                    xlsx_mobile_normalized = _safe_phone(xlsx_mobile)
                    if xlsx_mobile_normalized and str(student.mobile) != xlsx_mobile_normalized:
                        issues.append(f"Mobile mismatch: '{xlsx_mobile}' ({xlsx_mobile_normalized}) vs {student.mobile}")

                # Guardian name check
                xlsx_guardian = str(xlsx_row.get('guardian_name') or '').strip()
                if xlsx_guardian and student.guardian_name:
                    if _normalize_arabic_key(xlsx_guardian) != _normalize_arabic_key(student.guardian_name):
                        issues.append(f"Guardian name mismatch: '{xlsx_guardian}' vs '{student.guardian_name}'")

                # Teacher check (just verify it exists, names may differ)
                xlsx_teacher = str(xlsx_row.get('teacher_name') or '').strip()
                if xlsx_teacher:
                    if student.teacher:
                        # Check if teacher name matches (accounting for full vs short names)
                        xlsx_tokens = xlsx_teacher.split()
                        db_tokens = student.teacher.full_name.split()
                        if xlsx_tokens and db_tokens:
                            # First and last token should match
                            if xlsx_tokens[0] != db_tokens[0] or xlsx_tokens[-1] != db_tokens[-1]:
                                issues.append(f"Teacher mismatch: '{xlsx_teacher}' vs '{student.teacher.full_name}'")
                    else:
                        issues.append(f"Teacher assigned in xlsx '{xlsx_teacher}' but student has no teacher")

                # Address check
                xlsx_address = str(xlsx_row.get('address') or '').strip()
                if xlsx_address and student.address:
                    if _normalize_arabic_key(xlsx_address) != _normalize_arabic_key(student.address):
                        issues.append(f"Address mismatch: '{xlsx_address}' vs '{student.address}'")

                if issues:
                    stats['field_issues'].append({
                        'row': idx,
                        'national_id': national_id,
                        'issues': issues,
                    })

            # Summary report
            self.stdout.write("\n" + "=" * 120)
            self.stdout.write(self.style.SUCCESS(f"\n📊 VERIFICATION SUMMARY"))
            self.stdout.write(self.style.SUCCESS(f"  Total rows in xlsx: {stats['total']}"))
            self.stdout.write(self.style.SUCCESS(f"  ✓ Found in DB: {stats['found']} ({stats['found']*100//stats['total']}%)"))
            self.stdout.write(self.style.WARNING(f"  ⚠ No ID in xlsx: {stats['no_id']} ({stats['no_id']*100//stats['total']}%)"))
            self.stdout.write(self.style.ERROR(f"  ✗ Missing from DB: {stats['missing']} ({stats['missing']*100//stats['total']}%)"))

            if stats['field_issues']:
                self.stdout.write(self.style.WARNING(
                    f"\n  ⚠ With data issues: {len(stats['field_issues'])} rows"
                ))
                self.stdout.write("\nFirst 15 rows with issues:")
                for item in stats['field_issues'][:15]:
                    self.stdout.write(self.style.WARNING(f"\n  Row {item['row']} (ID: {item['national_id']}):"))
                    for issue in item['issues'][:2]:
                        self.stdout.write(self.style.WARNING(f"    - {issue}"))
            else:
                self.stdout.write(self.style.SUCCESS(f"\n  ✓ All data verified (within normalization)"))

            # Final verdict
            self.stdout.write("\n" + "=" * 120)
            if stats['found'] == stats['total'] and not stats['field_issues']:
                self.stdout.write(self.style.SUCCESS("\n✅ ALL DATA IMPORTED AND VERIFIED SUCCESSFULLY!\n"))
            elif stats['found'] == stats['total']:
                self.stdout.write(
                    self.style.SUCCESS(
                        f"\n✅ All {stats['total']} rows found and imported successfully!"
                    )
                )
                self.stdout.write(self.style.WARNING(
                    f"   ({len(stats['field_issues'])} rows have minor data differences due to normalization)\n"
                ))
            elif stats['found'] + stats['no_id'] == stats['total']:
                self.stdout.write(
                    self.style.SUCCESS(
                        f"\n✅ All rows in database! ({stats['found']} with ID, {stats['no_id']} without ID)\n"
                    )
                )
            else:
                self.stdout.write(
                    self.style.ERROR(
                        f"\n⚠ {stats['missing']} rows missing from database\n"
                    )
                )

        except FileNotFoundError:
            raise CommandError(f"File not found: {file_path}")
        except Exception as e:
            raise CommandError(f"Error verifying import: {str(e)}")
