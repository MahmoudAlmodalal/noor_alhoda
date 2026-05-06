"""
Management command to import student data from an Excel file.
Usage: python manage.py import_excel_data /path/to/file.xlsx [--verify]
"""
import openpyxl
from datetime import datetime, date
from django.core.management.base import BaseCommand, CommandError

from accounts.models import User
from students.services.excel_import_service import excel_bulk_import


def _verify_import(rows, stdout, style):
    from students.models import Student

    matched = 0
    missing = []
    mismatches = []

    for row in rows:
        raw_nid = str(row.get("national_id") or "").strip()
        name = str(row.get("full_name") or "").strip()

        student = (
            Student.objects
            .select_related("user", "teacher")
            .filter(user__national_id=raw_nid)
            .first()
        ) if raw_nid else None

        if student is None and name:
            # Synthetic IDs: normalize_row() built a 97-prefix id — look up by name
            student = Student.objects.select_related("user", "teacher").filter(
                full_name=name
            ).first()

        if student is None:
            missing.append(name or raw_nid or "—")
            continue

        matched += 1

        # Field checks
        if name and student.full_name != name:
            mismatches.append(
                f"  [{raw_nid or name}] الاسم: DB={student.full_name!r}  xlsx={name!r}"
            )

    stdout.write(style.SUCCESS(f"\n--- نتائج التحقق ---"))
    stdout.write(style.SUCCESS(f"✓ مطابق في قاعدة البيانات: {matched} / {len(rows)}"))

    if missing:
        stdout.write(style.ERROR(f"✗ غير موجود في DB ({len(missing)}):"))
        for label in missing[:20]:
            stdout.write(style.ERROR(f"  {label}"))

    if mismatches:
        stdout.write(style.WARNING(f"⚠ اختلاف في الاسم ({len(mismatches)}):"))
        for msg in mismatches[:20]:
            stdout.write(style.WARNING(msg))

    if not missing and not mismatches:
        stdout.write(style.SUCCESS("✓ جميع البيانات صحيحة!"))


class Command(BaseCommand):
    help = "Import student data from an Excel file."

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str, help="Path to the Excel file")
        parser.add_argument(
            "--verify",
            action="store_true",
            default=False,
            help="After import, cross-check each row against the database.",
        )

    def handle(self, *args, **options):
        file_path = options["file_path"]

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = None

            # Look for the sheet with data (skip empty sheets)
            for sheet_name in workbook.sheetnames:
                candidate_sheet = workbook[sheet_name]
                if candidate_sheet.max_row > 1:
                    sheet = candidate_sheet
                    break

            if not sheet:
                raise CommandError("No sheet with data found in Excel file.")

            # Find header row (look for 'الاسم رباعي' which is the student name column)
            header_row = None
            for row_idx in range(1, min(10, sheet.max_row + 1)):
                for cell in sheet[row_idx]:
                    if cell.value and 'اسم' in str(cell.value):
                        header_row = row_idx
                        break
                if header_row:
                    break

            if not header_row:
                raise CommandError("Could not find header row in Excel file.")

            # Get headers from the identified row
            headers = []
            for cell in sheet[header_row]:
                headers.append(cell.value)

            if not headers or not any(headers):
                raise CommandError("Excel file headers are empty.")

            # Map columns by position (not by name since names repeat)
            # Student info: cols 1-8
            # Empty space: cols 9-13
            # Address/Health/Skills/Courses: cols 14-17
            # Guardian info: cols 18-20
            # Bank info: cols 21-23
            # Teacher/Affiliation: cols 24-25

            column_index_map = {
                1: 'sequence_number',
                2: 'full_name',
                3: 'national_id',
                4: 'birthdate',
                5: 'age',
                6: 'grade',
                7: 'mobile',
                8: 'whatsapp',
                14: 'address',
                15: 'health_status',
                16: 'skills',
                17: 'previous_courses',
                18: 'guardian_name',
                19: 'guardian_national_id',
                20: 'guardian_mobile',
                21: 'bank_account_number',
                22: 'bank_account_name',
                23: 'bank_account_type',
                24: 'teacher_name',
                25: 'affiliation',
            }

            # Read rows starting from the row after headers
            rows = []
            for row_cells in sheet.iter_rows(min_row=header_row + 1, values_only=False):
                # Skip completely empty rows
                if not any(cell.value for cell in row_cells):
                    continue

                row_dict = {}
                for col_idx, cell in enumerate(row_cells):
                    if col_idx in column_index_map:
                        field_name = column_index_map[col_idx]
                        value = cell.value

                        # Convert datetime objects to date strings
                        if field_name == 'birthdate' and isinstance(value, datetime):
                            value = value.date()
                        if isinstance(value, date) and not isinstance(value, datetime):
                            value = value.isoformat()

                        row_dict[field_name] = value

                # Only add rows that have at least a name or ID
                if row_dict.get('full_name') or row_dict.get('national_id'):
                    rows.append(row_dict)

            if not rows:
                raise CommandError("Excel file has no data rows.")

            # Get or create admin user (default admin for importing)
            admin_user = User.objects.filter(role="admin").first()
            if not admin_user:
                self.stdout.write(
                    self.style.WARNING(
                        "No admin user found. Creating one with phone 0599100001..."
                    )
                )
                from accounts.services.auth_services import user_register

                admin_user = user_register(
                    phone_number="0599100001",
                    national_id="0000000000",
                    password="admin123",
                    role="admin",
                )

            # Import the data
            result = excel_bulk_import(creator=admin_user, rows=rows)

            self.stdout.write(
                self.style.SUCCESS(f"✓ Created: {result['created_count']} students")
            )
            self.stdout.write(
                self.style.SUCCESS(f"✓ Updated: {result['updated_count']} students")
            )

            if result["errors"]:
                self.stdout.write(
                    self.style.WARNING(
                        f"\n⚠ {result['error_count']} errors encountered:"
                    )
                )
                for error in result["errors"]:
                    self.stdout.write(
                        self.style.WARNING(
                            f"  Row {error['row']}: {error['message']}"
                        )
                    )
            else:
                self.stdout.write(self.style.SUCCESS("✓ No errors!"))

            # Verify import if requested
            if options.get("verify"):
                _verify_import(rows, self.stdout, self.style)

        except FileNotFoundError:
            raise CommandError(f"File not found: {file_path}")
        except Exception as e:
            raise CommandError(f"Error importing Excel file: {str(e)}")
