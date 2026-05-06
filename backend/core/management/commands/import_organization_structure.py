"""
Management command to import organization structure (teachers/staff) from Excel file.
Usage: python manage.py import_organization_structure /path/to/file.xlsx
"""
import openpyxl
import hashlib
from datetime import datetime, date
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from accounts.models import User
from teacher.models import Teacher


class Command(BaseCommand):
    help = "Import organization structure (teachers/staff) from an Excel file."

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str, help="Path to the Excel file")

    def handle(self, *args, **options):
        file_path = options["file_path"]

        try:
            workbook = openpyxl.load_workbook(file_path)

            # Find the sheet with data
            sheet = None
            for sheet_name in workbook.sheetnames:
                candidate_sheet = workbook[sheet_name]
                if candidate_sheet.max_row > 3:
                    sheet = candidate_sheet
                    break

            if not sheet:
                raise CommandError("No sheet with data found in Excel file.")

            # Find header row (look for 'الإسم')
            header_row = None
            for row_idx in range(1, min(10, sheet.max_row + 1)):
                for cell in sheet[row_idx]:
                    if cell.value and 'الإسم' in str(cell.value):
                        header_row = row_idx
                        break
                if header_row:
                    break

            if not header_row:
                raise CommandError("Could not find header row in Excel file.")

            # Map columns by enumerate index (0-based)
            # Actual Excel layout: A(empty), B(#), C(name), D(id), E(phone),
            # F(empty), G(wallet_name), H(wallet_number), I(student_count),
            # J(marital_status), K(empty), L(empty), M(family_count), N(job_title)
            column_index_map = {
                2: 'full_name',                     # Col C: الإسم
                3: 'national_id',                   # Col D: رقم الهوية
                4: 'contact_number',                # Col E: رقو التواصل
                6: 'wallet_name',                   # Col G: اسم المحفظة
                7: 'wallet_number',                 # Col H: رقم المحفظة
                9: 'marital_status',                # Col J: الحالة
                12: 'family_members_count',         # Col M: عدد أفراد الأسرة
                13: 'job_title',                    # Col N: المسمى الوظيفي
            }

            # Get or create admin user
            admin_user = User.objects.filter(role="admin").first()
            if not admin_user:
                self.stdout.write(
                    self.style.WARNING(
                        "No admin user found. Please create one first."
                    )
                )
                return

            # Read rows and create teachers
            created_count = 0
            updated_count = 0
            errors = []
            processed_ids = set()  # Track IDs to handle duplicates

            with transaction.atomic():
                for row_cells in sheet.iter_rows(min_row=header_row + 1, values_only=False):
                    # Skip completely empty rows
                    if not any(cell.value for cell in row_cells):
                        continue

                    row_dict = {}
                    for col_idx, cell in enumerate(row_cells):
                        if col_idx in column_index_map:
                            field_name = column_index_map[col_idx]
                            value = cell.value

                            # Convert datetime objects to date strings
                            if field_name == 'birthdate':
                                if isinstance(value, datetime):
                                    value = value.date().isoformat()
                                elif isinstance(value, date):
                                    value = value.isoformat()
                                elif isinstance(value, str) and value:
                                    # Handle DD\MM\YYYY or DD/MM/YYYY format
                                    try:
                                        value = value.replace('\\', '/')
                                        parts = value.split('/')
                                        if len(parts) == 3:
                                            # Assume DD/MM/YYYY format
                                            value = f"{parts[2]}-{parts[1]}-{parts[0]}"
                                    except:
                                        value = None
                            elif isinstance(value, date) and not isinstance(value, datetime):
                                value = value.isoformat()

                            row_dict[field_name] = value

                    # Convert family_members_count to int if present
                    if row_dict.get('family_members_count'):
                        try:
                            row_dict['family_members_count'] = int(row_dict['family_members_count'])
                        except (ValueError, TypeError):
                            row_dict['family_members_count'] = None

                    # Map marital status from Arabic to choice values
                    marital_mapping = {
                        'متزوج': 'married',
                        'أعزب': 'single',
                    }
                    if row_dict.get('marital_status'):
                        row_dict['marital_status'] = marital_mapping.get(
                            str(row_dict['marital_status']).strip(),
                            ''
                        )

                    # Map job title from Arabic to choice values
                    job_mapping = {
                        'محفظ': 'teacher',
                        'محفظ استقبال': 'teacher_reception',
                        'محفظ حلقة سنة': 'teacher_year_circle',
                        'محفظ حلقة منتدى': 'teacher_forum_circle',
                        'مساعد محفظ': 'teacher_assistant',
                        'معلم دورات': 'course_instructor',
                        'مساعد إداري + محفظ': 'admin_teacher',
                        'مدير المركز': 'teacher',  # Default to teacher
                        'نائب المدير': 'teacher',  # Default to teacher
                    }
                    if row_dict.get('job_title'):
                        row_dict['job_title'] = job_mapping.get(
                            str(row_dict['job_title']).strip(),
                            'teacher'
                        )

                    # Only process rows with a name or national ID
                    if not (row_dict.get('full_name') or row_dict.get('national_id')):
                        continue

                    try:
                        full_name = str(row_dict.get('full_name') or '').strip()
                        phone_number = str(row_dict.get('contact_number') or '').strip()
                        national_id = row_dict.get('national_id')

                        # Convert national_id to string and strip, or generate one if missing
                        if national_id:
                            national_id = str(national_id).strip()
                        if not national_id:
                            # Generate a synthetic ID from the name (hash-based)
                            name_hash = hashlib.md5(full_name.encode()).hexdigest()[:10]
                            national_id = name_hash

                        # Handle duplicate national_ids by appending phone number
                        if national_id in processed_ids:
                            # Duplicate ID in this import - make it unique
                            national_id = f"{national_id}_{phone_number[-4:]}"

                        processed_ids.add(national_id)

                        # Create or update user for teacher
                        user, created = User.objects.get_or_create(
                            national_id=national_id,
                            defaults={
                                'phone_number': phone_number,
                                'role': 'teacher',
                                'is_active': True,
                            }
                        )

                        if not created:
                            user.phone_number = phone_number
                            user.save()

                        # Prepare teacher defaults and updates
                        teacher_data = {
                            'full_name': full_name,
                            'ring_name': row_dict.get('wallet_name') or full_name,  # Use wallet_name as ring_name
                            'wallet_name': row_dict.get('wallet_name') or '',
                            'wallet_number': row_dict.get('wallet_number') or '',
                            'education_qualification': '',  # Not in Excel file
                            'last_tajweed_course': '',  # Not in Excel file
                            'marital_status': row_dict.get('marital_status') or '',
                            'family_members_count': row_dict.get('family_members_count'),
                            'job_title': row_dict.get('job_title') or 'teacher',
                        }

                        # Create or update teacher profile
                        teacher, t_created = Teacher.objects.get_or_create(
                            user=user,
                            defaults=teacher_data
                        )

                        if not t_created:
                            for field, value in teacher_data.items():
                                if value is not None and value != '':
                                    setattr(teacher, field, value)
                            teacher.save()

                        if t_created:
                            created_count += 1
                        else:
                            updated_count += 1

                    except Exception as e:
                        errors.append({
                            "name": row_dict.get('full_name', 'Unknown'),
                            "message": str(e)
                        })

            self.stdout.write(
                self.style.SUCCESS(f"✓ Created: {created_count} teachers")
            )
            self.stdout.write(
                self.style.SUCCESS(f"✓ Updated: {updated_count} teachers")
            )

            if errors:
                self.stdout.write(
                    self.style.WARNING(f"\n⚠ {len(errors)} errors encountered:")
                )
                for error in errors[:10]:  # Show first 10 errors
                    self.stdout.write(
                        self.style.WARNING(
                            f"  {error['name']}: {error['message']}"
                        )
                    )
            else:
                self.stdout.write(self.style.SUCCESS("✓ No errors!"))

        except FileNotFoundError:
            raise CommandError(f"File not found: {file_path}")
        except Exception as e:
            raise CommandError(f"Error importing Excel file: {str(e)}")
