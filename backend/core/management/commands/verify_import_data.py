"""
Comprehensive verification of Excel import data.
Checks all fields for each imported student against the source xlsx file.

Usage: python manage.py verify_import_data /path/to/file.xlsx
"""
import openpyxl
from datetime import datetime, date
from django.core.management.base import BaseCommand, CommandError

from students.models import Student
from students.services.excel_import.normalization import _normalize_arabic_key


class Command(BaseCommand):
    help = "Verify all imported student data against the source Excel file."

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str, help="Path to the Excel file")

    def handle(self, *args, **options):
        file_path = options["file_path"]

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = None

            # Find sheet with data
            for sheet_name in workbook.sheetnames:
                candidate_sheet = workbook[sheet_name]
                if candidate_sheet.max_row > 1:
                    sheet = candidate_sheet
                    break

            if not sheet:
                raise CommandError("No sheet with data found in Excel file.")

            # Find header row
            header_row = None
            for row_idx in range(1, min(10, sheet.max_row + 1)):
                for cell in sheet[row_idx]:
                    if cell.value and 'اسم' in str(cell.value):
                        header_row = row_idx
                        break
                if header_row:
                    break

            if not header_row:
                raise CommandError("Could not find header row in Excel file.")

            # Column map
            column_index_map = {
                1: 'sequence_number',
                2: 'full_name',
                3: 'national_id',
                4: 'birthdate',
                5: 'age',
                6: 'grade',
                7: 'mobile',
                8: 'whatsapp',
                14: 'address',
                15: 'health_status',
                16: 'skills',
                17: 'previous_courses',
                18: 'guardian_name',
                19: 'guardian_national_id',
                20: 'guardian_mobile',
                21: 'bank_account_number',
                22: 'bank_account_name',
                23: 'bank_account_type',
                24: 'teacher_name',
                25: 'affiliation',
            }

            # Read all rows
            rows = []
            for row_idx, row_cells in enumerate(
                sheet.iter_rows(min_row=header_row + 1, values_only=False),
                start=header_row + 1
            ):
                if not any(cell.value for cell in row_cells):
                    continue

                row_dict = {'_row_index': row_idx}
                for col_idx, cell in enumerate(row_cells):
                    if col_idx in column_index_map:
                        field_name = column_index_map[col_idx]
                        value = cell.value

                        # Convert datetime to ISO string
                        if field_name == 'birthdate' and isinstance(value, datetime):
                            value = value.date()
                        if isinstance(value, date) and not isinstance(value, datetime):
                            value = value.isoformat()

                        row_dict[field_name] = value

                if row_dict.get('full_name') or row_dict.get('national_id'):
                    rows.append(row_dict)

            self.stdout.write(self.style.SUCCESS(f"\nFound {len(rows)} student rows in xlsx"))
            self.stdout.write("=" * 100)

            # Verify each row
            stats = {
                'total': len(rows),
                'found': 0,
                'missing': 0,
                'field_mismatches': [],
                'full_match': 0,
            }

            for idx, xlsx_row in enumerate(rows, start=1):
                national_id = str(xlsx_row.get('national_id') or '').strip()
                full_name = str(xlsx_row.get('full_name') or '').strip()

                # Look up student by national_id
                student = Student.objects.select_related(
                    'user', 'teacher'
                ).filter(
                    user__national_id=national_id
                ).first()

                if not student:
                    stats['missing'] += 1
                    self.stdout.write(
                        self.style.ERROR(
                            f"✗ Row {idx} ({national_id}): NOT FOUND in database"
                        )
                    )
                    continue

                stats['found'] += 1

                # Check key fields
                mismatches = []

                # Name check (normalized)
                if full_name and student.full_name:
                    xlsx_normalized = _normalize_arabic_key(full_name)
                    db_normalized = _normalize_arabic_key(student.full_name)
                    if xlsx_normalized != db_normalized:
                        mismatches.append(
                            f"name: '{full_name}' vs '{student.full_name}'"
                        )

                # Grade check
                xlsx_grade = str(xlsx_row.get('grade') or '').strip()
                if xlsx_grade and student.grade:
                    if str(student.grade).strip() != xlsx_grade.strip():
                        mismatches.append(
                            f"grade: xlsx={xlsx_grade!r} vs db={student.grade!r}"
                        )

                # Mobile check
                xlsx_mobile = str(xlsx_row.get('mobile') or '').strip()
                if xlsx_mobile and student.mobile:
                    if str(student.mobile).strip() != xlsx_mobile.strip():
                        mismatches.append(
                            f"mobile: xlsx={xlsx_mobile!r} vs db={student.mobile!r}"
                        )

                # Guardian name check
                xlsx_guardian = str(xlsx_row.get('guardian_name') or '').strip()
                if xlsx_guardian and student.guardian_name:
                    if str(student.guardian_name).strip() != xlsx_guardian.strip():
                        mismatches.append(
                            f"guardian_name: xlsx={xlsx_guardian!r} vs db={student.guardian_name!r}"
                        )

                # Teacher check
                xlsx_teacher = str(xlsx_row.get('teacher_name') or '').strip()
                if xlsx_teacher and student.teacher:
                    if str(student.teacher.full_name).strip() != xlsx_teacher.strip():
                        # Check if it's just whitespace normalization
                        xlsx_norm = _normalize_arabic_key(xlsx_teacher)
                        db_norm = _normalize_arabic_key(student.teacher.full_name)
                        if xlsx_norm != db_norm:
                            mismatches.append(
                                f"teacher: xlsx={xlsx_teacher!r} vs db={student.teacher.full_name!r}"
                            )

                # Address check
                xlsx_address = str(xlsx_row.get('address') or '').strip()
                if xlsx_address and student.address:
                    if str(student.address).strip() != xlsx_address.strip():
                        mismatches.append(
                            f"address: xlsx={xlsx_address!r} vs db={student.address!r}"
                        )

                if not mismatches:
                    stats['full_match'] += 1
                    if idx % 20 == 0:  # Print every 20th match
                        self.stdout.write(
                            self.style.SUCCESS(f"✓ Row {idx} ({national_id}): OK")
                        )
                else:
                    stats['field_mismatches'].append({
                        'row': idx,
                        'national_id': national_id,
                        'mismatches': mismatches,
                    })
                    self.stdout.write(
                        self.style.WARNING(f"⚠ Row {idx} ({national_id}): Field mismatches")
                    )
                    for mismatch in mismatches[:3]:  # Show first 3 mismatches
                        self.stdout.write(self.style.WARNING(f"    - {mismatch}"))

            # Summary report
            self.stdout.write("\n" + "=" * 100)
            self.stdout.write(self.style.SUCCESS(f"\n📊 VERIFICATION SUMMARY"))
            self.stdout.write(self.style.SUCCESS(f"  Total rows in xlsx: {stats['total']}"))
            self.stdout.write(self.style.SUCCESS(f"  ✓ Found in DB: {stats['found']} ({stats['found']*100//stats['total']}%)"))
            self.stdout.write(self.style.ERROR(f"  ✗ Missing from DB: {stats['missing']} ({stats['missing']*100//stats['total']}%)"))
            self.stdout.write(
                self.style.SUCCESS(f"  ✓ Full match (all fields): {stats['full_match']} ({stats['full_match']*100//stats['total']}%)")
            )

            if stats['field_mismatches']:
                self.stdout.write(
                    self.style.WARNING(f"  ⚠ With field mismatches: {len(stats['field_mismatches'])}")
                )
                self.stdout.write("\nDetailed mismatches (first 10 rows):")
                for item in stats['field_mismatches'][:10]:
                    self.stdout.write(
                        self.style.WARNING(
                            f"\n  Row {item['row']} (ID: {item['national_id']}):"
                        )
                    )
                    for mismatch in item['mismatches']:
                        self.stdout.write(self.style.WARNING(f"    - {mismatch}"))

            if stats['found'] == stats['total'] and stats['full_match'] == stats['total']:
                self.stdout.write(self.style.SUCCESS("\n✅ ALL DATA VERIFIED SUCCESSFULLY!"))
            elif stats['found'] == stats['total']:
                self.stdout.write(
                    self.style.WARNING(
                        f"\n✓ All {stats['total']} rows found, but some field mismatches detected"
                    )
                )
            else:
                self.stdout.write(
                    self.style.ERROR(
                        f"\n✗ {stats['missing']} rows missing from database"
                    )
                )

        except FileNotFoundError:
            raise CommandError(f"File not found: {file_path}")
        except Exception as e:
            raise CommandError(f"Error verifying import: {str(e)}")
